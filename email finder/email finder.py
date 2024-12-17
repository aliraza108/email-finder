import requests
from bs4 import BeautifulSoup
import re
from urllib.parse import urljoin, urlparse
import openpyxl as xl
from openpyxl import load_workbook
import warnings
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import deque

# Suppress the XMLParsedAsHTMLWarning
warnings.filterwarnings("ignore", category=UserWarning, module="bs4")

# Regular expression pattern to match email addresses
email_pattern = re.compile(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}')

# Set up a session to reuse the connection and improve performance
session = requests.Session()

# Add headers to simulate a real browser request
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
}
session.headers.update(headers)

# List of priority pages to search for emails
priority_pages = [
    "contact", "about", "write-for-us", "contact-us", "privacy-policy", "contact.html", "about-us"
]

# Function to fetch emails from a webpage
def fetch_emails_from_page(url):
    try:
        response = session.get(url, timeout=3)  # Reduced timeout for faster responses
        response.raise_for_status()  # Check for HTTP errors
        content = response.text

        # Find all email addresses in the content
        emails = set(email_pattern.findall(content))

        # Also search within bold text elements
        soup = BeautifulSoup(content, 'html.parser')
        for bold_text in soup.find_all(['b', 'strong']):
            emails.update(email_pattern.findall(bold_text.get_text()))

        return emails  # Return unique email addresses
    except requests.RequestException as e:
        print(f"Error fetching {url}: {e}")
        return set()

# Function to find all internal links on a webpage (limit to top-level links)
def find_internal_links(url, base_url, max_links=5):
    try:
        response = session.get(url, timeout=3)
        response.raise_for_status()

        # Use lxml parser if available
        soup = BeautifulSoup(response.text, 'lxml')

        internal_links = set()

        # Find all anchor tags with href attributes
        for a_tag in soup.find_all('a', href=True):
            link = a_tag['href']
            full_link = urljoin(base_url, link)
            if urlparse(full_link).netloc == urlparse(base_url).netloc:
                internal_links.add(full_link)
                if len(internal_links) >= max_links:
                    break  # Limit the number of links to crawl to avoid overloading

        return internal_links
    except requests.RequestException as e:
        print(f"Error fetching {url}: {e}")
        return set()

# Function to find priority links on the base domain
def find_priority_links(base_url):
    try:
        response = session.get(base_url, timeout=3)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')

        priority_links = set()
        for a_tag in soup.find_all('a', href=True):
            link = a_tag['href']
            full_link = urljoin(base_url, link)

            # Check if the link contains any of the priority page keywords
            if any(page in full_link.lower() for page in priority_pages):
                priority_links.add(full_link)

        return priority_links
    except requests.RequestException as e:
        print(f"Error fetching priority links from {base_url}: {e}")
        return set()

# Function to process emails from multiple pages concurrently
def process_domain_emails(domain):
    base_url = f"https://{domain.strip()}"
    to_visit = deque([base_url])  # Start with the base URL
    visited = set()
    emails_found = set()

    # Find priority links on the base domain
    priority_links = find_priority_links(base_url)
    to_visit.extend(priority_links)

    # Perform the breadth-first search (BFS) to gather internal links and emails
    with ThreadPoolExecutor(max_workers=15) as executor:  # Increased worker count for concurrency
        while to_visit:
            current_url = to_visit.popleft()
            if current_url not in visited:
                visited.add(current_url)
                future = executor.submit(fetch_emails_from_page, current_url)
                try:
                    emails = future.result()
                    emails_found.update(emails)

                    # Find new internal links from the current page if no emails are found
                    if not emails_found:
                        new_links = find_internal_links(current_url, base_url, max_links=3)
                        to_visit.extend(new_links - visited)

                except Exception as e:
                    print(f"Error processing {current_url}: {e}")

                if emails_found:
                    break  # Exit early if emails are found

    return emails_found

# Main function for processing the Excel file
def emailfinder():
    file_path = 'email finder/file.xlsx'  # Use the correct path to your file
    wb = load_workbook(file_path)
    ws = wb.active

    # Iterate through rows in the first column to get website domains
    for row in ws.iter_rows(min_col=1, max_col=1, min_row=2):  # Skip header row
        for cell in row:
            domain = cell.value
            if domain:  # Check if domain is not empty
                emails_found = process_domain_emails(domain)

                # Output the result for the domain
                if emails_found:
                    emails_string = "; ".join(emails_found)  # Join multiple emails with a semicolon
                    print(f"Emails found for {domain}: {emails_string}")

                    # Save the found emails to column G of the corresponding row
                    ws[f'G{cell.row}'] = emails_string
                else:
                    print(f"No emails found for {domain}")
                    ws[f'G{cell.row}'] = ''

    # Save the updated Excel file
    wb.save(file_path)

# Run the emailfinder function
emailfinder()
