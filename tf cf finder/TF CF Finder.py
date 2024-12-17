import openpyxl as xl
import requests
import json
import time
import re 
API_KEY = 'FDEDE810FE13F605A8512B20B45E3028'
file_path = 'email finder/file.xlsx'
wb = xl.load_workbook(file_path)
ws = wb.active

valid_suffixes = [".com", ".uk", ".org", ".edu", ".eu", ".co", ".net"]

def fetch_metrics_for_url(url):
    api_url = f'https://api.majestic.com/api/json?app_api_key={API_KEY}&cmd=GetIndexItemInfo&items=1&item0={url}&datasource=fresh'
    try:
        response = requests.get(api_url)
        json_data = response.json()
        
        if 'DataTables' in json_data and 'Results' in json_data['DataTables']:
            results = json_data['DataTables']['Results'].get('Data', [])
            if results:
                result = results[0]
                tf = result.get('TrustFlow', 'TF not found')
                cf = result.get('CitationFlow', 'CF not found')
                return tf, cf
        return 'Data not found', 'Data not found'
    except Exception as e:
        return f'Error: {str(e)}', f'Error: {str(e)}'

for row in range(2, ws.max_row + 1):
    start = time.time()
    url = ws.cell(row=row, column=1).value 
    if url:
        for a in valid_suffixes:
            if(a in url):
                tf, cf = fetch_metrics_for_url(url)
                ws.cell(row=row, column=2).value = tf  
                ws.cell(row=row, column=3).value = cf  
            else:
                end = time.time()
                print("The time of execution of above program is :",
        (end-start) * 10**3, "ms")

wb.save('email finder/updated_file.xlsx')
